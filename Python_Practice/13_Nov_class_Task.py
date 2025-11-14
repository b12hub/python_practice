import os
from abc import ABC, abstractmethod
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from time import strftime
from openpyxl.styles import Font

OUTPUT_DIR = "invoices"
timestamp = strftime("%Y-%m-%d %H:%M:%S")
if not os.path.exists(OUTPUT_DIR):
    os.makedirs(OUTPUT_DIR)

class InvoiceGenerator(ABC):
    def __init__(self, client_name, items):
        self.client_name = client_name
        self.items = items

    @abstractmethod
    def calculate_total(self):
        pass

    @abstractmethod
    def generate_invoice(self):
        pass

# ---------- PDF Generator with Table ----------
class PDFInvoiceGenerator(InvoiceGenerator):
    def calculate_total(self):
        return sum(self.items.values())

    def generate_invoice(self):
        file_path = os.path.join(OUTPUT_DIR, "invoice.pdf")
        doc = SimpleDocTemplate(file_path, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        elements.append(Paragraph(f"Invoice for {self.client_name}", styles['Title']))

        data = [["Item", "Price ($)"]]
        for item, price in self.items.items():
            data.append([item, f"{price:.2f}"])
        data.append(["Total", f"{self.calculate_total():.2f}"])
        data.append(["Generated at", timestamp])

        table = Table(data, hAlign='LEFT')
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
            ('ALIGN',(0,0),(-1,-1),'LEFT'),
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('BACKGROUND',(0,1),(-1,-1),colors.beige)
        ]))
        elements.append(table)
        doc.build(elements)
        return file_path

class ExcelInvoiceGenerator( InvoiceGenerator):
    def calculate_total(self):
        return sum(self.items.values())

    def generate_invoice(self):
        file_path = os.path.join(OUTPUT_DIR, f"Invoice-{timestamp}.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice"

        # Header row
        ws.append(["Item", "Price ($)", "Client Name"])
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Add items
        for item, price in self.items.items():
            ws.append([item, price, self.client_name])

        # Total row
        ws.append(["Total", self.calculate_total(), self.client_name])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

        # Timestamp row
        ws.append(["Generated at", timestamp, ""])
        for cell in ws[ws.max_row]:
            cell.font = Font(italic=True)

        wb.save(file_path)
        return file_path

class HTMLInvoiceGenerator(InvoiceGenerator):
    def calculate_total(self):
        return sum(self.items.values())

    def generate_invoice(self):
        file_path = os.path.join(OUTPUT_DIR, "invoice.html")
        html_content = f"""
        <!DOCTYPE html>
        <html lang='en'>
        <head>
            <meta charset='UTF-8'>
            <title>Invoice for {self.client_name}</title>
            <style>
                table {{border-collapse: collapse; width: 50%;}}
                th, td {{border: 1px solid black; padding: 8px; text-align: left;}}
                th {{background-color: #4CAF50; color: white;}}
                tr:nth-child(even){{background-color: #f2f2f2;}}
            </style>
        </head>
        <body>
            <h1>Invoice for {self.client_name}</h1>
            <table>
                <tr><th>Item</th><th>Price ($)</th></tr>
        """
        for item, price in self.items.items():
            html_content += f"<tr><td>{item}</td><td>{price:.2f}</td></tr>"
        html_content += f"<tr><td>Total</td><td>{self.calculate_total():.2f}</td></tr>"
        html_content += f"<tr><td colspan='2'>Generated at: {timestamp}</td></tr>"
        html_content += "</table></body></html>"

        with open(file_path, "w") as f:
            f.write(html_content)
        return file_path

class InvoiceManager:
    def __init__(self):
        self.invoice_generators = []

    def add_generator(self, generator: InvoiceGenerator):
        self.invoice_generators.append(generator)

    def export_invoice(self):
        for generator in self.invoice_generators:
            generator.generate_invoice()


if __name__ == "__main__":
    invoice_manager = InvoiceManager()
    items = {"Apples": 1.5, "Bananas": 2.0}
    invoice_manager.add_generator(PDFInvoiceGenerator("John Doe", items))
    invoice_manager.add_generator(ExcelInvoiceGenerator("John Doe", items))
    invoice_manager.add_generator(HTMLInvoiceGenerator("John Doe", items))
    invoice_manager.export_invoice()
